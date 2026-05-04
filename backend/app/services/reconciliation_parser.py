"""
对账文件解析服务

支持 CSV 和 Excel 格式，提取运费账单数据。
"""
import io
import csv
import openpyxl
from datetime import datetime
from typing import Iterator, Any


def validate_file(file_content: bytes, file_name: str) -> tuple[bool, str]:
    """
    验证上传文件格式和大小。
    返回 (is_valid, error_message)
    """
    # 大小限制: 10MB
    MAX_SIZE = 10 * 1024 * 1024
    if len(file_content) > MAX_SIZE:
        return False, f"文件大小超过限制 (最大 10MB), 当前: {len(file_content) / 1024 / 1024:.2f}MB"

    # 格式验证
    lower_name = file_name.lower()
    if lower_name.endswith('.csv'):
        return True, ""
    elif lower_name.endswith(('.xlsx', '.xls')):
        return True, ""
    else:
        return False, f"不支持的文件格式: {file_name}, 仅支持 CSV 和 Excel (.xlsx/.xls)"


def _convert_value(value: Any) -> Any:
    """转换单元格值，处理 None 和空字符串"""
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        if stripped == '':
            return None
        return stripped
    return value


def _parse_float(value: Any) -> float | None:
    """解析浮点数"""
    if value is None:
        return None
    try:
        if isinstance(value, str):
            value = value.replace(',', '').replace('$', '').strip()
        return float(value)
    except (ValueError, TypeError):
        return None


def _parse_int(value: Any) -> int | None:
    """解析整数"""
    if value is None:
        return None
    try:
        if isinstance(value, str):
            value = value.replace(',', '').strip()
        return int(float(value))
    except (ValueError, TypeError):
        return None


def parse_excel_shipping_sheet(file_content: bytes) -> Iterator[dict]:
    """
    解析 Excel 运费账单 sheet。
    支持 zbhome_2026_3_账单update.xlsx 的格式。
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)

    # 优先查找 "运费" sheet
    sheet_name = None
    for name in wb.sheetnames:
        if '运费' in name or 'shipping' in name.lower():
            sheet_name = name
            break

    # 如果没找到，使用第一个 sheet
    if sheet_name is None:
        sheet_name = wb.sheetnames[0]

    ws = wb[sheet_name]

    # 查找表头行（通常在第2行）
    header_row = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True)):
        if row and any(h in str(h) for h in row if h):
            # 检查是否包含关键字段
            row_strs = [str(h).lower() if h else '' for h in row]
            if any('跟踪号' in s or 'tracking' in s or '总价' in s or 'total' in s for s in row_strs):
                header_row = i + 1
                break

    if header_row is None:
        header_row = 2  # 默认第2行

    # 读取表头
    headers = []
    for cell in ws[header_row]:
        headers.append(_convert_value(cell.value))

    # 建立列映射
    col_map = {}
    for i, h in enumerate(headers):
        if h:
            col_map[h] = i

    # 遍历数据行
    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True)):
        # 跳过空行
        if not row or all(v is None for v in row):
            continue

        def _safe_col(key: str):
            if key not in col_map:
                return None
            idx = col_map[key]
            if idx >= len(row):
                return None
            return row[idx]

        # 提取关键字段
        record = {
            'row_no': row_idx + header_row + 1,
            'ship_date': _convert_value(_safe_col('打单日期')),
            'carrier': _convert_value(_safe_col('物流商')),
            'service': _convert_value(_safe_col('物流服务')),
            'sku': _convert_value(_safe_col('SKU')),
            'tracking_no': _convert_value(_safe_col('跟踪号')),
            'total_amount': _parse_float(_safe_col('总价')),
            'base_amount': _parse_float(_safe_col('基础价')),
            'qty': _parse_int(_safe_col('数量')),
            'length_in': _parse_float(_safe_col('长(in)')),
            'width_in': _parse_float(_safe_col('宽(in)')),
            'height_in': _parse_float(_safe_col('高(in)')),
            'weight_lb': _parse_float(_safe_col('重量(lbs)')),
            'dim_weight_lb': _parse_float(_safe_col('体积重(lbs)')),
            'billed_weight': _parse_float(_safe_col('计费重')),
            'weight_unit': _convert_value(_safe_col('重量单位')),
            'warehouse': _convert_value(_safe_col('发货仓')),
            'zip_code': _convert_value(_safe_col('目的地邮编')),
            'zone': _convert_value(_safe_col('区')),
            'fuel': _parse_float(_safe_col('Fuel')),
            'ahs_dim': _parse_float(_safe_col('AHS-DIM')),
            'ahs_weight': _parse_float(_safe_col('AHS-Weight')),
            'ahs_packaging': _parse_float(_safe_col('AHS-Packaging')),
            'oversize': _parse_float(_safe_col('Oversize')),
            'non_mach': _parse_float(_safe_col('Non-Mach Surcharge')),
            'signature': _parse_float(_safe_col('Signature')),
            'adc': _parse_float(_safe_col('ADC')),
            'das': _parse_float(_safe_col('DAS')),
            'extend': _parse_float(_safe_col('EXTEND')),
            'das_remote': _parse_float(_safe_col('DAS-Remote')),
            'unauthorized': _parse_float(_safe_col('Unauthorized')),
            'residential': _parse_float(_safe_col('Residential')),
            'store': _convert_value(_safe_col('店铺')),
            'remark1': _convert_value(_safe_col('备注一')),
            'remark2': _convert_value(_safe_col('备注二')),
            'remark3': _convert_value(_safe_col('备注三')),
            'fee_type': _convert_value(_safe_col('运费类型')),
            'order_type': _convert_value(_safe_col('订单类型')),
        }

        # 过滤掉没有跟踪号的记录（可能是空行或汇总行）
        if record['tracking_no']:
            yield record


def parse_csv_shipping_file(file_content: bytes) -> Iterator[dict]:
    """解析 CSV 运费账单文件"""
    text = file_content.decode('utf-8-sig')
    reader = csv.DictReader(text.splitlines())

    for row_idx, row in enumerate(reader, start=2):
        record = {
            'row_no': row_idx,
            'ship_date': _convert_value(row.get('打单日期')),
            'carrier': _convert_value(row.get('物流商')),
            'service': _convert_value(row.get('物流服务')),
            'sku': _convert_value(row.get('SKU')),
            'tracking_no': _convert_value(row.get('跟踪号')),
            'total_amount': _parse_float(row.get('总价')),
            'base_amount': _parse_float(row.get('基础价')),
            'qty': _parse_int(row.get('数量')),
            'length_in': _parse_float(row.get('长(in)')),
            'width_in': _parse_float(row.get('宽(in)')),
            'height_in': _parse_float(row.get('高(in)')),
            'weight_lb': _parse_float(row.get('重量(lbs)')),
            'dim_weight_lb': _parse_float(row.get('体积重(lbs)')),
            'billed_weight': _parse_float(row.get('计费重')),
            'weight_unit': _convert_value(row.get('重量单位')),
            'warehouse': _convert_value(row.get('发货仓')),
            'zip_code': _convert_value(row.get('目的地邮编')),
            'zone': _convert_value(row.get('区')),
            'fuel': _parse_float(row.get('Fuel')),
            'ahs_dim': _parse_float(row.get('AHS-DIM')),
            'ahs_weight': _parse_float(row.get('AHS-Weight')),
            'ahs_packaging': _parse_float(row.get('AHS-Packaging')),
            'oversize': _parse_float(row.get('Oversize')),
            'non_mach': _parse_float(row.get('Non-Mach Surcharge')),
            'signature': _parse_float(row.get('Signature')),
            'adc': _parse_float(row.get('ADC')),
            'das': _parse_float(row.get('DAS')),
            'extend': _parse_float(row.get('EXTEND')),
            'das_remote': _parse_float(row.get('DAS-Remote')),
            'unauthorized': _parse_float(row.get('Unauthorized')),
            'residential': _parse_float(row.get('Residential')),
            'store': _convert_value(row.get('店铺')),
            'remark1': _convert_value(row.get('备注一')),
            'remark2': _convert_value(row.get('备注二')),
            'remark3': _convert_value(row.get('备注三')),
            'fee_type': _convert_value(row.get('运费类型')),
            'order_type': _convert_value(row.get('订单类型')),
        }

        if record['tracking_no']:
            yield record


def parse_shipping_file(file_content: bytes, file_name: str) -> Iterator[dict]:
    """
    根据文件类型解析运费账单。
    返回每条记录的 dict。
    """
    lower_name = file_name.lower()
    if lower_name.endswith('.csv'):
        return parse_csv_shipping_file(file_content)
    else:
        return parse_excel_shipping_sheet(file_content)
