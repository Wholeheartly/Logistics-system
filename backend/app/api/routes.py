"""
FastAPI 路由
"""
from fastapi import FastAPI, Depends, HTTPException, UploadFile, File, Query, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse
from pydantic import BaseModel, Field, validator
from sqlalchemy.orm import Session
from sqlalchemy.exc import IntegrityError, SQLAlchemyError
from typing import Optional
import math
import uuid
import os
import shutil
import re
import logging
from datetime import datetime
from pathlib import Path

logger = logging.getLogger(__name__)

from app.utils.db import get_session
from app.services.product_service import (
    get_product_by_sku, search_products,
    create_product, parse_excel_to_products, bulk_create_products,
    validate_product_data, check_duplicate_sku,
    convert_all_products, get_product_with_converted,
)
from app.services.comparison_service import compare_all_carriers
from app.services.reconciliation_service import reconcile_batch
from app.services.reconciliation_parser import validate_file, parse_shipping_file
from app.services.reconciliation_engine import run_reconciliation
from app.services.config_discovery_service import (
    sync_config_items, apply_config_change,
    discover_carrier_configs, discover_surcharge_configs, discover_system_configs
)
from app.services.auth_service import (
    create_user, authenticate_user, create_access_token, get_current_user,
    has_permission, approve_user, disable_user, enable_user, list_users,
    create_password_reset_token, reset_password,
    hash_password, verify_password
)
from app.models.models import (
    ReconciliationBatch, ReconciliationDetail,
    ConfigItem, ConfigHistory, ConfigAuditLog, ConfigCategory,
    User, UserRole, UserStatus, ROLE_PERMISSIONS, DEPARTMENT_ROLE_MAP,
    Permission, ROLE_DISPLAY_NAMES, PERMISSION_DISPLAY_NAMES
)

app = FastAPI(title="跨境电商物流比价系统")

# CORS配置：生产环境应通过环境变量限制来源，开发环境可保持宽松
import os
_cors_origins_env = os.environ.get("CORS_ALLOWED_ORIGINS", "")
if _cors_origins_env.strip():
    # 生产环境：使用配置的白名单
    CORS_ALLOWED_ORIGINS = [origin.strip() for origin in _cors_origins_env.split(",") if origin.strip()]
else:
    # 开发环境：默认允许本地前端
    CORS_ALLOWED_ORIGINS = [
        "http://localhost:3000",
        "http://localhost:5173",
        "http://127.0.0.1:3000",
        "http://127.0.0.1:5173",
    ]

app.add_middleware(
    CORSMiddleware,
    allow_origins=CORS_ALLOWED_ORIGINS,
    allow_methods=["GET", "POST", "PUT", "DELETE", "OPTIONS"],
    allow_headers=["Authorization", "Content-Type", "Accept", "X-Requested-With"],
    allow_credentials=True,
    max_age=600,
)

# 创建头像上传目录
AVATAR_DIR = Path(__file__).parent.parent.parent / "uploads" / "avatars"
AVATAR_DIR.mkdir(parents=True, exist_ok=True)

# 挂载静态文件服务
app.mount("/uploads", StaticFiles(directory=str(AVATAR_DIR.parent)), name="uploads")


# ── Security Helpers ──

def _get_token_from_request(request: Request, token_query: Optional[str] = None) -> Optional[str]:
    """
    统一提取认证令牌。
    优先级：1. Authorization Header (Bearer)  2. URL Query参数（向后兼容，已弃用）
    """
    # 1. 优先从Header获取（安全方式）
    auth_header = request.headers.get("Authorization", "")
    if auth_header.startswith("Bearer "):
        return auth_header[7:].strip()

    # 2. 兼容Query参数（向后兼容，记录弃用警告）
    if token_query and token_query.strip():
        logger.warning(
            "Token passed via URL query parameter is deprecated and insecure. "
            "Use Authorization: Bearer <token> header instead."
        )
        return token_query.strip()

    return None


# ── Request / Response Models ──

class ShippingCompareRequest(BaseModel):
    sku: str = Field(..., description="产品 SKU")
    zip_code: str = Field(..., description="目的地邮编 (5位)")
    warehouse: str = Field(default="CA", description="发货仓 (CA/NJ/TX/SAV)")
    is_residential: bool = Field(default=False, description="是否住宅地址")


class ShippingZoneCompareRequest(BaseModel):
    sku: str = Field(..., description="产品 SKU")
    zip_code: str = Field(..., description="目的地邮编 (5位)")
    warehouses: list[str] = Field(default=["CA", "NJ", "TX", "SAV"], description="要比价的发货仓列表")
    is_residential: bool = Field(default=False, description="是否住宅地址")


class WarehouseZoneResult(BaseModel):
    warehouse: str = Field(..., description="发货仓代码")
    warehouse_name: str = Field(..., description="发货仓名称")
    zone: int = Field(..., description="Zone编号")
    carriers: list[dict] = Field(default=[], description="该仓库各渠道运费结果")


class OrderItem(BaseModel):
    order_no: str = Field(default="")
    sku: str = Field(default="")
    total_goods_price: Optional[float] = None
    freight: Optional[float] = None
    tax: Optional[float] = 0.0
    discount: Optional[float] = 0.0
    platform_subsidy: Optional[float] = 0.0
    buyer_paid: Optional[float] = None
    platform_fee: Optional[float] = None


class ReconciliationRequest(BaseModel):
    orders: list[OrderItem]


class ProductSearchRequest(BaseModel):
    keyword: str = Field(default="")


class ProductCreateRequest(BaseModel):
    sku: str = Field(..., min_length=1, max_length=50, description="产品 SKU")
    name: str = Field(..., min_length=1, max_length=200, description="产品名称")
    model: Optional[str] = Field(default=None, max_length=100, description="型号")
    specification: Optional[str] = Field(default=None, max_length=500, description="规格")
    price: Optional[float] = Field(default=0.0, ge=0, description="价格")
    stock_quantity: Optional[int] = Field(default=0, ge=0, description="库存数量")
    length_cm: float = Field(..., gt=0, description="长度(cm)")
    width_cm: float = Field(..., gt=0, description="宽度(cm)")
    height_cm: float = Field(..., gt=0, description="高度(cm)")
    gross_weight_kg: float = Field(..., gt=0, description="毛重(kg)")
    category: Optional[str] = Field(default=None, max_length=100, description="分类")
    brand: Optional[str] = Field(default=None, max_length=100, description="品牌")
    supplier: Optional[str] = Field(default=None, max_length=200, description="供应商")
    description: Optional[str] = Field(default=None, description="描述")
    status: Optional[str] = Field(default="active", description="状态")


class ConfigUpdateRequest(BaseModel):
    new_value: str = Field(..., description="新值")
    reason: str = Field(default="", description="变更原因")


class LoginRequest(BaseModel):
    username: str = Field(..., min_length=3, max_length=50)
    password: str = Field(..., min_length=6)
    remember_me: bool = Field(default=False)


class RegisterRequest(BaseModel):
    username: str = Field(..., min_length=3, max_length=50)
    password: str = Field(..., min_length=6)
    email: Optional[str] = Field(default=None)
    display_name: Optional[str] = Field(default=None)
    phone: Optional[str] = Field(default=None)
    department: Optional[str] = Field(default=None)


class ForgotPasswordRequest(BaseModel):
    username: str = Field(..., min_length=3, max_length=50)


class PasswordResetRequest(BaseModel):
    token: str = Field(...)
    new_password: str = Field(..., min_length=6)


class UserCreateRequest(BaseModel):
    username: str = Field(..., min_length=3, max_length=50)
    password: str = Field(..., min_length=6)
    email: Optional[str] = Field(default=None)
    display_name: Optional[str] = Field(default=None)
    role: str = Field(default=UserRole.OPERATOR)
    phone: Optional[str] = Field(default=None)
    department: Optional[str] = Field(default=None)


class RoleAssignRequest(BaseModel):
    role: str = Field(..., description="目标角色")


class PermissionCheckRequest(BaseModel):
    permission: str = Field(..., description="要检查的权限")


# ── Routes ──

@app.get("/api/health")
def health():
    return {"status": "ok"}


@app.post("/api/shipping/compare")
def shipping_compare(req: ShippingCompareRequest):
    session = get_session()
    try:
        result = compare_all_carriers(
            session,
            sku=req.sku,
            zip_code=req.zip_code,
            warehouse=req.warehouse,
            is_residential=req.is_residential,
        )
        return result
    finally:
        session.close()


@app.post("/api/shipping/zone-compare")
def shipping_zone_compare(req: ShippingZoneCompareRequest):
    """
    多发货仓区域比价。
    对指定 SKU 和邮编，计算所有指定仓库的运费并横向对比。
    """
    from app.services.zone_service import get_zone

    session = get_session()
    try:
        product = session.query(Product).filter(Product.sku == req.sku).first()
        if not product:
            return {"error": f"未找到 SKU: {req.sku}"}

        warehouse_names = {
            "CA": "CA 西部仓",
            "NJ": "NJ 东部仓",
            "TX": "TX 南部仓",
            "SAV": "SAV 东南部仓",
            "SC": "SC 东南部仓",
        }

        warehouse_results = []
        for wh in req.warehouses:
            wh_upper = wh.upper()
            zone = get_zone(session, wh_upper, req.zip_code)
            if zone is None:
                continue

            result = compare_all_carriers(
                session,
                sku=req.sku,
                zip_code=req.zip_code,
                warehouse=wh_upper,
                is_residential=req.is_residential,
            )
            if result.get("error"):
                continue

            warehouse_results.append({
                "warehouse": wh_upper,
                "warehouse_name": warehouse_names.get(wh_upper, wh_upper),
                "zone": zone,
                "results": result.get("results", []),
                "errors": result.get("errors", []),
            })

        # 按总价升序排列所有仓库的所有渠道
        all_carriers = []
        for wh in warehouse_results:
            for r in wh["results"]:
                all_carriers.append({
                    "warehouse": wh["warehouse"],
                    "warehouse_name": wh["warehouse_name"],
                    "zone": wh["zone"],
                    **r,
                })

        all_carriers.sort(key=lambda x: x["total"])

        # 标记最便宜的（仅标记总价最低的一项）
        if all_carriers:
            all_carriers[0]["is_cheapest"] = True

        return {
            "sku": req.sku,
            "zip_code": req.zip_code,
            "is_residential": req.is_residential,
            "warehouses": warehouse_results,
            "all_carriers": all_carriers,
        }
    finally:
        session.close()


@app.post("/api/reconciliation/batch")
def reconciliation_batch(req: ReconciliationRequest):
    orders = [o.model_dump() for o in req.orders]
    result = reconcile_batch(orders)
    return result


# ========== 对账系统 API ==========

@app.post("/api/reconciliation/upload")
async def reconciliation_upload(
    request: Request,
    file: UploadFile = File(...),
    token: Optional[str] = Query(default=None),
):
    """上传对账文件并执行批量比对（仅财务和管理员）"""
    effective_token = _get_token_from_request(request, token)
    if not effective_token:
        raise HTTPException(status_code=401, detail="未提供认证令牌")
    session = get_session()
    try:
        current = get_current_user(session, effective_token)
        if not current or not has_permission(current, "reconciliation.upload"):
            raise HTTPException(status_code=403, detail="权限不足，仅财务和管理员可上传对账文件")
        content = await file.read()

        # 验证文件
        is_valid, error_msg = validate_file(content, file.filename)
        if not is_valid:
            raise HTTPException(status_code=400, detail=error_msg)

        # 创建批次记录
        batch = ReconciliationBatch(
            batch_no=f"RC{datetime.now().strftime('%Y%m%d%H%M%S')}{uuid.uuid4().hex[:6].upper()}",
            name=file.filename,
            file_name=file.filename,
            file_size=len(content),
            file_type=file.filename.split('.')[-1].lower(),
            status="processing",
            created_at=datetime.now(),
        )
        session.add(batch)
        session.commit()
        session.refresh(batch)

        # 解析文件
        try:
            records = parse_shipping_file(content, file.filename)
        except Exception as e:
            batch.status = "failed"
            batch.error_message = f"文件解析失败: {str(e)}"
            session.commit()
            raise HTTPException(status_code=400, detail=f"文件解析失败: {str(e)}")

        # 执行对账
        result = run_reconciliation(session, batch, records)

        return {
            "batch_id": batch.id,
            "batch_no": batch.batch_no,
            "status": "completed",
            "total": result['total'],
            "matched": result['matched'],
            "diff": result['diff'],
        }
    finally:
        session.close()


@app.get("/api/reconciliation/batches")
def reconciliation_batches(
    request: Request,
    token: Optional[str] = Query(default=None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
):
    """查询对账批次列表（仅财务和管理员）"""
    effective_token = _get_token_from_request(request, token)
    if not effective_token:
        raise HTTPException(status_code=401, detail="未提供认证令牌")
    session = get_session()
    try:
        current = get_current_user(session, effective_token)
        if not current or not has_permission(current, "reconciliation.view"):
            raise HTTPException(status_code=403, detail="权限不足，仅财务和管理员可查看对账")
        query = session.query(ReconciliationBatch).order_by(ReconciliationBatch.created_at.desc())
        total = query.count()
        batches = query.offset((page - 1) * page_size).limit(page_size).all()

        return {
            "total": total,
            "page": page,
            "page_size": page_size,
            "batches": [
                {
                    "id": b.id,
                    "batch_no": b.batch_no,
                    "name": b.name,
                    "file_name": b.file_name,
                    "total_records": b.total_records,
                    "matched_records": b.matched_records,
                    "diff_records": b.diff_records,
                    "status": b.status,
                    "created_at": b.created_at.isoformat() if b.created_at else None,
                    "completed_at": b.completed_at.isoformat() if b.completed_at else None,
                }
                for b in batches
            ],
        }
    finally:
        session.close()


@app.get("/api/reconciliation/batch/{batch_id}/details")
def reconciliation_details(
    request: Request,
    batch_id: int,
    token: Optional[str] = Query(default=None),
    has_diff: Optional[bool] = Query(None),
    diff_type: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=500),
):
    """查询对账明细（仅财务和管理员）"""
    import json
    effective_token = _get_token_from_request(request, token)
    if not effective_token:
        raise HTTPException(status_code=401, detail="未提供认证令牌")
    session = get_session()
    try:
        current = get_current_user(session, effective_token)
        if not current or not has_permission(current, "reconciliation.view"):
            raise HTTPException(status_code=403, detail="权限不足，仅财务和管理员可查看对账明细")
        query = session.query(ReconciliationDetail).filter(ReconciliationDetail.batch_id == batch_id)

        if has_diff is not None:
            query = query.filter(ReconciliationDetail.has_diff == has_diff)

        if diff_type:
            query = query.filter(ReconciliationDetail.diff_types.contains(diff_type))

        total = query.count()
        details = query.order_by(ReconciliationDetail.row_no).offset((page - 1) * page_size).limit(page_size).all()

        return {
            "total": total,
            "page": page,
            "page_size": page_size,
            "details": [
                {
                    "id": d.id,
                    "row_no": d.row_no,
                    "file_order_no": d.file_order_no,
                    "file_tracking_no": d.file_tracking_no,
                    "file_sku": d.file_sku,
                    "file_carrier": d.file_carrier,
                    "file_total_amount": d.file_total_amount,
                    "file_base_amount": d.file_base_amount,
                    "file_weight_lb": d.file_weight_lb,
                    "file_billed_weight": d.file_billed_weight,
                    "file_zone": d.file_zone,
                    "file_warehouse": d.file_warehouse,
                    "file_zip_code": d.file_zip_code,
                    "sys_total_amount": d.sys_total_amount,
                    "sys_base_amount": d.sys_base_amount,
                    "sys_weight_lb": d.sys_weight_lb,
                    "sys_billed_weight": d.sys_billed_weight,
                    "sys_zone": d.sys_zone,
                    "has_diff": d.has_diff,
                    "diff_types": json.loads(d.diff_types) if d.diff_types else [],
                    "diff_amount": d.diff_amount,
                    "diff_details": json.loads(d.diff_details) if d.diff_details else {},
                }
                for d in details
            ],
        }
    finally:
        session.close()


@app.get("/api/reconciliation/batch/{batch_id}/export")
def reconciliation_export(
    request: Request,
    batch_id: int,
    token: Optional[str] = Query(default=None),
):
    """导出差异报告为 Excel（仅财务和管理员）"""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import json

    effective_token = _get_token_from_request(request, token)
    if not effective_token:
        raise HTTPException(status_code=401, detail="未提供认证令牌")
    session = get_session()
    try:
        current = get_current_user(session, effective_token)
        if not current or not has_permission(current, "reconciliation.export"):
            raise HTTPException(status_code=403, detail="权限不足，仅财务和管理员可导出对账报告")
        batch = session.query(ReconciliationBatch).filter(ReconciliationBatch.id == batch_id).first()
        if not batch:
            raise HTTPException(status_code=404, detail="批次不存在")

        details = session.query(ReconciliationDetail).filter(
            ReconciliationDetail.batch_id == batch_id,
            ReconciliationDetail.has_diff == True
        ).order_by(ReconciliationDetail.row_no).all()

        # 创建 Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "差异报告"

        # 表头
        headers = [
            "行号", "订单号", "跟踪号", "SKU", "物流商", "服务",
            "文件总价", "系统总价", "差异金额",
            "文件基础价", "系统基础价",
            "文件重量", "系统重量",
            "文件计费重", "系统计费重",
            "文件Zone", "系统Zone",
            "差异类型", "差异详情",
        ]
        ws.append(headers)

        # 样式
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 数据
        for d in details:
            diff_types = json.loads(d.diff_types) if d.diff_types else []
            diff_details = json.loads(d.diff_details) if d.diff_details else {}

            diff_desc = []
            for dt in diff_types:
                type_names = {
                    "amount_mismatch": "金额不符",
                    "base_amount_mismatch": "基础运费不符",
                    "weight_mismatch": "重量不符",
                    "billed_weight_mismatch": "计费重不符",
                    "zone_mismatch": "Zone不符",
                    "surcharge_mismatch": "附加费不符",
                    "missing_in_system": "系统中缺失",
                    "carrier_mismatch": "物流商不匹配",
                }
                diff_desc.append(type_names.get(dt, dt))

            ws.append([
                d.row_no,
                d.file_order_no,
                d.file_tracking_no,
                d.file_sku,
                d.file_carrier,
                d.file_service,
                d.file_total_amount,
                d.sys_total_amount,
                d.diff_amount,
                d.file_base_amount,
                d.sys_base_amount,
                d.file_weight_lb,
                d.sys_weight_lb,
                d.file_billed_weight,
                d.sys_billed_weight,
                d.file_zone,
                d.sys_zone,
                ", ".join(diff_desc),
                json.dumps(diff_details, ensure_ascii=False),
            ])

        # 调整列宽
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        from fastapi.responses import StreamingResponse
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=diff_report_{batch.batch_no}.xlsx"}
        )
    finally:
        session.close()


# ========== 配置管理 API ==========

@app.post("/api/config/sync")
def config_sync():
    """同步系统配置项（扫描所有可配置字段）"""
    session = get_session()
    try:
        stats = sync_config_items(session)
        return {
            "success": True,
            "message": f"配置同步完成：新建 {stats['created']} 项，更新 {stats['updated']} 项，未变更 {stats['unchanged']} 项",
            "stats": stats,
        }
    finally:
        session.close()


@app.get("/api/config/items")
def config_items(
    category: Optional[str] = Query(None),
    keyword: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
):
    """查询配置项列表"""
    session = get_session()
    try:
        query = session.query(ConfigItem)

        if category:
            query = query.filter(ConfigItem.category == category)

        if keyword:
            query = query.filter(
                (ConfigItem.config_key.contains(keyword)) |
                (ConfigItem.display_name.contains(keyword)) |
                (ConfigItem.description.contains(keyword))
            )

        total = query.count()
        items = query.order_by(ConfigItem.category, ConfigItem.sub_category, ConfigItem.config_key) \
            .offset((page - 1) * page_size).limit(page_size).all()

        return {
            "total": total,
            "page": page,
            "page_size": page_size,
            "items": [
                {
                    "id": item.id,
                    "config_key": item.config_key,
                    "category": item.category,
                    "sub_category": item.sub_category,
                    "display_name": item.display_name,
                    "description": item.description,
                    "current_value": item.current_value,
                    "default_value": item.default_value,
                    "suggested_value": item.suggested_value,
                    "value_type": item.value_type,
                    "unit": item.unit,
                    "is_editable": item.is_editable,
                    "is_sensitive": item.is_sensitive,
                    "min_value": item.min_value,
                    "max_value": item.max_value,
                    "related_entity_type": item.related_entity_type,
                    "related_entity_id": item.related_entity_id,
                    "related_field": item.related_field,
                    "updated_at": item.updated_at.isoformat() if item.updated_at else None,
                }
                for item in items
            ],
        }
    finally:
        session.close()


@app.get("/api/config/categories")
def config_categories():
    """获取配置分类列表"""
    session = get_session()
    try:
        categories = session.query(ConfigItem.category).distinct().all()
        category_counts = {}
        for (cat,) in categories:
            count = session.query(ConfigItem).filter(ConfigItem.category == cat).count()
            category_counts[cat] = count

        category_names = {
            ConfigCategory.CARRIER: "物流商配置",
            ConfigCategory.SURCHARGE: "附加费配置",
            ConfigCategory.SYSTEM: "系统配置",
            ConfigCategory.ZONE: "Zone配置",
            ConfigCategory.RATE: "费率配置",
            ConfigCategory.PRODUCT: "产品配置",
        }

        return {
            "categories": [
                {
                    "key": cat,
                    "name": category_names.get(cat, cat),
                    "count": category_counts.get(cat, 0),
                }
                for cat in [c[0] for c in categories]
            ]
        }
    finally:
        session.close()


@app.get("/api/config/item/{item_id}")
def config_item_detail(item_id: int):
    """获取配置项详情（含历史记录）"""
    import json
    session = get_session()
    try:
        item = session.query(ConfigItem).filter(ConfigItem.id == item_id).first()
        if not item:
            raise HTTPException(status_code=404, detail="配置项不存在")

        histories = session.query(ConfigHistory).filter(ConfigHistory.config_id == item_id) \
            .order_by(ConfigHistory.created_at.desc()).limit(20).all()

        return {
            "id": item.id,
            "config_key": item.config_key,
            "category": item.category,
            "sub_category": item.sub_category,
            "display_name": item.display_name,
            "description": item.description,
            "current_value": item.current_value,
            "default_value": item.default_value,
            "suggested_value": item.suggested_value,
            "value_type": item.value_type,
            "unit": item.unit,
            "is_editable": item.is_editable,
            "is_sensitive": item.is_sensitive,
            "min_value": item.min_value,
            "max_value": item.max_value,
            "related_entity_type": item.related_entity_type,
            "related_entity_id": item.related_entity_id,
            "related_field": item.related_field,
            "created_at": item.created_at.isoformat() if item.created_at else None,
            "updated_at": item.updated_at.isoformat() if item.updated_at else None,
            "histories": [
                {
                    "id": h.id,
                    "old_value": h.old_value,
                    "new_value": h.new_value,
                    "change_reason": h.change_reason,
                    "changed_by": h.changed_by,
                    "change_type": h.change_type,
                    "created_at": h.created_at.isoformat() if h.created_at else None,
                }
                for h in histories
            ],
        }
    finally:
        session.close()


@app.post("/api/config/item/{item_id}/update")
def config_item_update(item_id: int, req: ConfigUpdateRequest):
    """更新配置项"""
    session = get_session()
    try:
        item = session.query(ConfigItem).filter(ConfigItem.id == item_id).first()
        if not item:
            raise HTTPException(status_code=404, detail="配置项不存在")

        if not item.is_editable:
            raise HTTPException(status_code=403, detail="该配置项不可编辑")

        # 验证数值范围（仅当配置项同时设置了最小值和最大值时）
        if item.value_type in ("float", "int") and item.min_value is not None and item.max_value is not None:
            try:
                val = float(req.new_value)
                if val < item.min_value or val > item.max_value:
                    raise HTTPException(
                        status_code=400,
                        detail=f"值必须在 {item.min_value} 到 {item.max_value} 之间"
                    )
            except ValueError:
                raise HTTPException(status_code=400, detail="无效的数值格式")

        # 应用变更
        apply_config_change(session, item, req.new_value, req.reason, "admin")

        # 记录审计日志
        audit = ConfigAuditLog(
            action="edit",
            config_key=item.config_key,
            config_id=item.id,
            details=f"旧值: {item.current_value} -> 新值: {req.new_value}",
            user_id="admin",
            user_name="管理员",
            created_at=datetime.now(),
        )
        session.add(audit)
        session.commit()

        return {
            "success": True,
            "message": "配置更新成功",
            "config_key": item.config_key,
            "new_value": req.new_value,
        }
    finally:
        session.close()


@app.post("/api/config/item/{item_id}/rollback/{history_id}")
def config_item_rollback(item_id: int, history_id: int):
    """回滚到历史版本"""
    session = get_session()
    try:
        item = session.query(ConfigItem).filter(ConfigItem.id == item_id).first()
        if not item:
            raise HTTPException(status_code=404, detail="配置项不存在")

        history = session.query(ConfigHistory).filter(
            ConfigHistory.id == history_id,
            ConfigHistory.config_id == item_id
        ).first()
        if not history:
            raise HTTPException(status_code=404, detail="历史记录不存在")

        # 应用回滚
        apply_config_change(session, item, history.old_value or item.default_value or "", "回滚到历史版本", "admin")

        # 记录审计日志
        audit = ConfigAuditLog(
            action="rollback",
            config_key=item.config_key,
            config_id=item.id,
            details=f"回滚到历史版本 #{history_id}",
            user_id="admin",
            user_name="管理员",
            created_at=datetime.now(),
        )
        session.add(audit)
        session.commit()

        return {
            "success": True,
            "message": "配置回滚成功",
            "config_key": item.config_key,
            "rolled_back_to": history.old_value,
        }
    finally:
        session.close()


@app.get("/api/config/audit-logs")
def config_audit_logs(
    config_key: Optional[str] = Query(None),
    action: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
):
    """查询审计日志"""
    session = get_session()
    try:
        query = session.query(ConfigAuditLog)

        if config_key:
            query = query.filter(ConfigAuditLog.config_key == config_key)
        if action:
            query = query.filter(ConfigAuditLog.action == action)

        total = query.count()
        logs = query.order_by(ConfigAuditLog.created_at.desc()) \
            .offset((page - 1) * page_size).limit(page_size).all()

        return {
            "total": total,
            "page": page,
            "page_size": page_size,
            "logs": [
                {
                    "id": log.id,
                    "action": log.action,
                    "config_key": log.config_key,
                    "details": log.details,
                    "user_id": log.user_id,
                    "user_name": log.user_name,
                    "ip_address": log.ip_address,
                    "created_at": log.created_at.isoformat() if log.created_at else None,
                }
                for log in logs
            ],
        }
    finally:
        session.close()


@app.get("/api/warehouses")
def list_warehouses():
    return {
        "warehouses": [
            {"code": "CA", "name": "西部仓 (默认)"},
            {"code": "NJ", "name": "东部仓"},
            {"code": "TX", "name": "南部仓"},
            {"code": "SAV", "name": "东南部仓"},
        ]
    }


# ========== 发货区域配置管理 API ==========

class ZoneMappingCreateRequest(BaseModel):
    warehouse: str = Field(..., min_length=1, max_length=10, description="发货仓代码")
    zip_prefix: str = Field(..., min_length=1, max_length=5, description="邮编前缀")
    zone: int = Field(..., ge=2, le=8, description="Zone编号 (2-8)")


class ZoneMappingUpdateRequest(BaseModel):
    zone: int = Field(..., ge=2, le=8, description="Zone编号 (2-8)")


@app.get("/api/zones")
def list_zones(
    request: Request,
    token: Optional[str] = Query(default=None),
    warehouse: Optional[str] = Query(None),
    zip_prefix: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=500),
):
    """查询发货区域配置列表（管理员）"""
    effective_token = _get_token_from_request(request, token)
    if not effective_token:
        raise HTTPException(status_code=401, detail="未提供认证令牌")
    session = get_session()
    try:
        current = get_current_user(session, effective_token)
        if not current or not has_permission(current, Permission.CONFIG_MANAGE):
            raise HTTPException(status_code=403, detail="权限不足，仅管理员可管理发货区域")

        query = session.query(ZoneMapping)

        if warehouse:
            query = query.filter(ZoneMapping.warehouse == warehouse.upper())
        if zip_prefix:
            query = query.filter(ZoneMapping.zip_prefix == zip_prefix)

        total = query.count()
        zones = query.order_by(ZoneMapping.warehouse, ZoneMapping.zip_prefix) \
            .offset((page - 1) * page_size).limit(page_size).all()

        return {
            "total": total,
            "page": page,
            "page_size": page_size,
            "zones": [
                {
                    "id": z.id,
                    "warehouse": z.warehouse,
                    "zip_prefix": z.zip_prefix,
                    "zone": z.zone,
                }
                for z in zones
            ],
        }
    finally:
        session.close()


@app.post("/api/zones")
def create_zone(
    request: Request,
    req: ZoneMappingCreateRequest,
    token: Optional[str] = Query(default=None),
):
    """新增发货区域配置（管理员）"""
    effective_token = _get_token_from_request(request, token)
    if not effective_token:
        raise HTTPException(status_code=401, detail="未提供认证令牌")
    session = get_session()
    try:
        current = get_current_user(session, effective_token)
        if not current or not has_permission(current, Permission.CONFIG_MANAGE):
            raise HTTPException(status_code=403, detail="权限不足")

        # 检查是否已存在
        existing = session.query(ZoneMapping).filter(
            ZoneMapping.warehouse == req.warehouse.upper(),
            ZoneMapping.zip_prefix == req.zip_prefix,
        ).first()
        if existing:
            raise HTTPException(status_code=400, detail="该仓库和邮编前缀的组合已存在")

        zone = ZoneMapping(
            warehouse=req.warehouse.upper(),
            zip_prefix=req.zip_prefix,
            zone=req.zone,
        )
        session.add(zone)
        session.commit()
        session.refresh(zone)

        return {
            "success": True,
            "message": "发货区域配置添加成功",
            "zone": {
                "id": zone.id,
                "warehouse": zone.warehouse,
                "zip_prefix": zone.zip_prefix,
                "zone": zone.zone,
            },
        }
    finally:
        session.close()


@app.post("/api/zones/{zone_id}/update")
def update_zone(
    request: Request,
    zone_id: int,
    req: ZoneMappingUpdateRequest,
    token: Optional[str] = Query(default=None),
):
    """更新发货区域配置（管理员）"""
    effective_token = _get_token_from_request(request, token)
    if not effective_token:
        raise HTTPException(status_code=401, detail="未提供认证令牌")
    session = get_session()
    try:
        current = get_current_user(session, effective_token)
        if not current or not has_permission(current, Permission.CONFIG_MANAGE):
            raise HTTPException(status_code=403, detail="权限不足")

        zone = session.query(ZoneMapping).filter(ZoneMapping.id == zone_id).first()
        if not zone:
            raise HTTPException(status_code=404, detail="发货区域配置不存在")

        zone.zone = req.zone
        session.commit()

        return {
            "success": True,
            "message": "发货区域配置更新成功",
            "zone": {
                "id": zone.id,
                "warehouse": zone.warehouse,
                "zip_prefix": zone.zip_prefix,
                "zone": zone.zone,
            },
        }
    finally:
        session.close()


@app.post("/api/zones/{zone_id}/delete")
def delete_zone(
    request: Request,
    zone_id: int,
    token: Optional[str] = Query(default=None),
):
    """删除发货区域配置（管理员）"""
    effective_token = _get_token_from_request(request, token)
    if not effective_token:
        raise HTTPException(status_code=401, detail="未提供认证令牌")
    session = get_session()
    try:
        current = get_current_user(session, effective_token)
        if not current or not has_permission(current, Permission.CONFIG_MANAGE):
            raise HTTPException(status_code=403, detail="权限不足")

        zone = session.query(ZoneMapping).filter(ZoneMapping.id == zone_id).first()
        if not zone:
            raise HTTPException(status_code=404, detail="发货区域配置不存在")

        session.delete(zone)
        session.commit()

        return {
            "success": True,
            "message": "发货区域配置删除成功",
        }
    finally:
        session.close()


@app.get("/api/zones/warehouses")
def list_zone_warehouses(
    request: Request,
    token: Optional[str] = Query(default=None),
):
    """获取所有有区域配置的仓库列表（管理员）"""
    effective_token = _get_token_from_request(request, token)
    if not effective_token:
        raise HTTPException(status_code=401, detail="未提供认证令牌")
    session = get_session()
    try:
        current = get_current_user(session, effective_token)
        if not current or not has_permission(current, Permission.CONFIG_MANAGE):
            raise HTTPException(status_code=403, detail="权限不足")

        warehouses = session.query(ZoneMapping.warehouse).distinct().all()
        return {
            "warehouses": [{"code": w[0], "name": w[0]} for w in warehouses],
        }
    finally:
        session.close()


# ========== 认证与授权 API ==========

@app.post("/api/auth/register")
def auth_register(req: RegisterRequest):
    """用户注册（需管理员审核）"""
    session = get_session()
    try:
        # 检查用户名是否已存在
        existing = session.query(User).filter(User.username == req.username).first()
        if existing:
            raise HTTPException(status_code=400, detail="用户名已存在")

        user = create_user(
            session,
            username=req.username,
            password=req.password,
            email=req.email,
            display_name=req.display_name,
            phone=req.phone,
            department=req.department,
        )
        return {
            "success": True,
            "message": "注册成功，请等待管理员审核",
            "user_id": user.id,
            "username": user.username,
            "status": user.status,
        }
    finally:
        session.close()


@app.post("/api/auth/login")
def auth_login(req: LoginRequest):
    """用户登录"""
    session = get_session()
    try:
        user, error = authenticate_user(session, req.username, req.password)
        if not user:
            raise HTTPException(status_code=401, detail=error)

        token = create_access_token(user)
        return {
            "success": True,
            "token": token,
            "user": {
                "id": user.id,
                "username": user.username,
                "display_name": user.display_name,
                "role": user.role,
                "status": user.status,
            },
            "permissions": ROLE_PERMISSIONS.get(user.role, []),
        }
    finally:
        session.close()


@app.get("/api/auth/me")
def auth_me(token: Optional[str] = Query(default=None)):
    """获取当前登录用户信息"""
    if not token or not token.strip():
        raise HTTPException(status_code=401, detail="未登录或令牌为空")
    session = get_session()
    try:
        user = get_current_user(session, token)
        if not user:
            raise HTTPException(status_code=401, detail="登录已过期或无效")
        return {
            "id": user.id,
            "username": user.username,
            "display_name": user.display_name,
            "email": user.email,
            "phone": user.phone,
            "role": user.role,
            "status": user.status,
            "department": user.department,
            "avatar_url": getattr(user, "avatar_url", None),
            "permissions": ROLE_PERMISSIONS.get(user.role, []),
        }
    finally:
        session.close()


@app.post("/api/auth/forgot-password")
def auth_forgot_password(req: ForgotPasswordRequest = None):
    """申请密码重置（支持 JSON body 和 Query 参数）"""
    if req is None:
        raise HTTPException(status_code=422, detail="请求体不能为空")
    session = get_session()
    try:
        user = session.query(User).filter(User.username == req.username).first()
        if not user:
            raise HTTPException(status_code=404, detail="用户不存在")

        token = create_password_reset_token(session, user.id)
        return {
            "success": True,
            "message": "密码重置链接已生成（演示模式：请使用以下令牌）",
            "token": token,
        }
    finally:
        session.close()


@app.post("/api/auth/reset-password")
def auth_reset_password(req: PasswordResetRequest = None):
    """重置密码"""
    if req is None:
        raise HTTPException(status_code=422, detail="请求体不能为空")
    # 密码强度校验
    pwd = req.new_password
    if len(pwd) < 6:
        raise HTTPException(status_code=400, detail="密码长度至少6位")
    if not any(c.isalpha() for c in pwd):
        raise HTTPException(status_code=400, detail="密码需包含字母")
    if not any(c.isdigit() for c in pwd):
        raise HTTPException(status_code=400, detail="密码需包含数字")

    session = get_session()
    try:
        success, message = reset_password(session, req.token, req.new_password)
        if not success:
            raise HTTPException(status_code=400, detail=message)
        return {"success": True, "message": message}
    finally:
        session.close()


# ========== 用户管理 API（管理员） ==========

@app.get("/api/users")
def users_list(
    token: str = Query(...),
    status: Optional[str] = Query(None),
    role: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    keyword: Optional[str] = Query(None),
):
    """查询用户列表（管理员）"""
    session = get_session()
    try:
        current = get_current_user(session, token)
        if not current or not has_permission(current, "user.manage"):
            raise HTTPException(status_code=403, detail="权限不足")
        
        # 查询用户
        query = session.query(User)
        
        # 状态筛选
        if status:
            query = query.filter(User.status == status)
        
        # 角色筛选
        if role:
            query = query.filter(User.role == role)
        
        # 关键词搜索
        if keyword:
            keyword_lower = keyword.lower()
            query = query.filter(
                (User.username.ilike(f'%{keyword_lower}%')) |
                (User.display_name.ilike(f'%{keyword_lower}%')) |
                (User.email.ilike(f'%{keyword_lower}%'))
            )
        
        # 计算总数
        total = query.count()
        
        # 分页
        users = query.order_by(User.created_at.desc()) \
            .offset((page - 1) * page_size) \
            .limit(page_size) \
            .all()
        
        return {
            "total": total,
            "page": page,
            "page_size": page_size,
            "users": [
                {
                    "id": u.id,
                    "username": u.username,
                    "display_name": u.display_name,
                    "email": u.email,
                    "role": u.role,
                    "status": u.status,
                    "phone": u.phone,
                    "department": u.department,
                    "avatar_url": getattr(u, "avatar_url", None),
                    "last_login_at": u.last_login_at.isoformat() if u.last_login_at else None,
                    "created_at": u.created_at.isoformat() if u.created_at else None,
                }
                for u in users
            ]
        }
    finally:
        session.close()


@app.post("/api/users")
def users_create(token: str = Query(...), req: UserCreateRequest = None):
    """管理员创建用户（带部门角色校验）"""
    if req is None:
        raise HTTPException(status_code=400, detail="请求体不能为空")
    session = get_session()
    try:
        current = get_current_user(session, token)
        if not current or not has_permission(current, "user.create"):
            raise HTTPException(status_code=403, detail="权限不足")

        # 校验部门与角色对应关系
        if req.department and req.department in DEPARTMENT_ROLE_MAP:
            allowed_roles = DEPARTMENT_ROLE_MAP[req.department]
            if req.role not in allowed_roles:
                raise HTTPException(
                    status_code=400,
                    detail=f"部门 '{req.department}' 不允许分配角色 '{req.role}'，允许的角色: {', '.join(allowed_roles)}"
                )

        existing = session.query(User).filter(User.username == req.username).first()
        if existing:
            raise HTTPException(status_code=400, detail="用户名已存在")

        user = create_user(
            session,
            username=req.username,
            password=req.password,
            email=req.email,
            display_name=req.display_name,
            role=req.role,
            phone=req.phone,
            department=req.department,
            created_by=current.id,
        )
        # 管理员创建的用户直接激活
        user.status = UserStatus.ACTIVE
        session.commit()

        return {
            "success": True,
            "message": "用户创建成功",
            "user_id": user.id,
        }
    finally:
        session.close()


@app.post("/api/users/{user_id}/approve")
def users_approve(user_id: int, token: str = Query(...)):
    """审核通过用户"""
    session = get_session()
    try:
        current = get_current_user(session, token)
        if not current or not has_permission(current, "user.approve"):
            raise HTTPException(status_code=403, detail="权限不足")

        success, message = approve_user(session, user_id, current.id)
        if not success:
            raise HTTPException(status_code=400, detail=message)
        return {"success": True, "message": message}
    finally:
        session.close()


@app.post("/api/users/{user_id}/disable")
def users_disable(user_id: int, token: str = Query(...)):
    """禁用用户"""
    session = get_session()
    try:
        current = get_current_user(session, token)
        if not current or not has_permission(current, "user.delete"):
            raise HTTPException(status_code=403, detail="权限不足")

        success, message = disable_user(session, user_id)
        if not success:
            raise HTTPException(status_code=400, detail=message)
        return {"success": True, "message": message}
    finally:
        session.close()


@app.post("/api/users/{user_id}/enable")
def users_enable(user_id: int, token: str = Query(...)):
    """重新启用被禁用的用户（需要 user.edit 权限）"""
    session = get_session()
    try:
        current = get_current_user(session, token)
        if not current or not has_permission(current, "user.edit"):
            raise HTTPException(status_code=403, detail="权限不足，无法执行用户启用操作")

        success, message = enable_user(session, user_id, current)
        if not success:
            raise HTTPException(status_code=400, detail=message)
        return {"success": True, "message": message}
    finally:
        session.close()


@app.post("/api/users/{user_id}/reset-password")
def admin_reset_password(user_id: int, token: str = Query(...), new_password: str = Query(...)):
    """管理员重置用户密码"""
    session = get_session()
    try:
        current = get_current_user(session, token)
        if not current or not has_permission(current, "user.edit"):
            raise HTTPException(status_code=403, detail="权限不足")

        user = session.query(User).filter(User.id == user_id).first()
        if not user:
            raise HTTPException(status_code=404, detail="用户不存在")

        user.password_hash = hash_password(new_password)
        user.updated_at = datetime.now()
        session.commit()
        return {"success": True, "message": "密码重置成功"}
    finally:
        session.close()


class UserUpdateRequest(BaseModel):
    role: Optional[str] = Field(None)
    status: Optional[str] = Field(None)
    department: Optional[str] = Field(None)


@app.post("/api/users/{user_id}/update")
def users_update(user_id: int, token: str = Query(...), req: UserUpdateRequest = None):
    """更新用户信息（管理员）"""
    if req is None:
        raise HTTPException(status_code=400, detail="请求体不能为空")
    session = get_session()
    try:
        current = get_current_user(session, token)
        if not current or not has_permission(current, "user.edit"):
            raise HTTPException(status_code=403, detail="权限不足")
        
        # 不能修改自己
        if current.id == user_id and req.role is not None:
            raise HTTPException(status_code=400, detail="不能修改自己的角色")
        
        user = session.query(User).filter(User.id == user_id).first()
        if not user:
            raise HTTPException(status_code=404, detail="用户不存在")
        
        # 更新角色
        if req.role is not None:
            if req.role not in ['admin', 'operator', 'finance']:
                raise HTTPException(status_code=400, detail="无效的角色")
            
            # 验证部门与角色对应关系
            if req.department and req.department in DEPARTMENT_ROLE_MAP:
                allowed_roles = DEPARTMENT_ROLE_MAP[req.department]
                if req.role not in allowed_roles:
                    raise HTTPException(
                        status_code=400,
                        detail=f"部门 '{req.department}' 不允许分配角色 '{req.role}'"
                    )
            
            user.role = req.role
        
        # 更新状态
        if req.status is not None:
            if req.status not in ['pending', 'active', 'disabled']:
                raise HTTPException(status_code=400, detail="无效的状态")
            user.status = req.status
        
        # 更新部门
        if req.department is not None:
            # 验证部门与角色对应关系
            if req.department in DEPARTMENT_ROLE_MAP:
                allowed_roles = DEPARTMENT_ROLE_MAP[req.department]
                if user.role not in allowed_roles:
                    raise HTTPException(
                        status_code=400,
                        detail=f"部门 '{req.department}' 不允许当前角色 '{user.role}'"
                    )
            user.department = req.department
        
        user.updated_at = datetime.now()
        session.commit()
        
        return {"success": True, "message": "用户更新成功"}
    finally:
        session.close()


# ========== 个人信息管理 API ==========

class ProfileUpdateRequest(BaseModel):
    display_name: Optional[str] = Field(default=None)
    email: Optional[str] = Field(default=None)
    phone: Optional[str] = Field(default=None)
    department: Optional[str] = Field(default=None)
    avatar_url: Optional[str] = Field(default=None)

    @validator('email')
    def validate_email(cls, v):
        if v is not None and v.strip() != '':
            if not re.match(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$', v):
                raise ValueError('邮箱格式不正确')
        return v

    @validator('phone')
    def validate_phone(cls, v):
        if v is not None and v.strip() != '':
            if not re.match(r'^[\d\s\-+()]{5,20}$', v):
                raise ValueError('手机号格式不正确')
        return v

    @validator('display_name')
    def validate_display_name(cls, v):
        if v is not None and v.strip() != '':
            if len(v.strip()) > 100:
                raise ValueError('显示名称不能超过100个字符')
        return v


@app.get("/api/users/departments")
def list_departments(token: str = Query(...)):
    """获取部门与角色对应关系"""
    session = get_session()
    try:
        current = get_current_user(session, token)
        if not current:
            raise HTTPException(status_code=401, detail="未登录")
        return {
            "departments": [
                {"name": dept, "allowed_roles": roles}
                for dept, roles in DEPARTMENT_ROLE_MAP.items()
            ]
        }
    finally:
        session.close()


@app.get("/api/profile")
def get_profile(request: Request, token: Optional[str] = Query(default=None)):
    """获取当前用户个人信息"""
    effective_token = _get_token_from_request(request, token)
    if not effective_token:
        raise HTTPException(status_code=401, detail="未登录或令牌为空")
    session = get_session()
    try:
        current = get_current_user(session, effective_token)
        if not current:
            raise HTTPException(status_code=401, detail="未登录或会话已过期")
        return {
            "id": current.id,
            "username": current.username,
            "display_name": current.display_name,
            "email": current.email,
            "phone": current.phone,
            "department": current.department,
            "role": current.role,
            "status": current.status,
            "avatar_url": getattr(current, "avatar_url", None),
            "permissions": ROLE_PERMISSIONS.get(current.role, []),
        }
    finally:
        session.close()


@app.post("/api/profile")
def update_profile(request: Request, token: Optional[str] = Query(default=None), req: ProfileUpdateRequest = None):
    """更新当前用户个人信息"""
    effective_token = _get_token_from_request(request, token)
    if not effective_token:
        raise HTTPException(status_code=401, detail="未登录或令牌为空")
    if req is None:
        raise HTTPException(status_code=400, detail="请求体不能为空")
    session = get_session()
    try:
        current = get_current_user(session, effective_token)
        if not current:
            raise HTTPException(status_code=401, detail="未登录或会话已过期")

        # 如果修改部门，校验新部门是否允许当前角色
        if req.department and req.department in DEPARTMENT_ROLE_MAP:
            allowed_roles = DEPARTMENT_ROLE_MAP[req.department]
            if current.role not in allowed_roles:
                raise HTTPException(
                    status_code=400,
                    detail=f"部门 '{req.department}' 不允许当前角色 '{current.role}'"
                )

        # 更新字段：空字符串视为 None，避免触发数据库唯一约束
        if req.display_name is not None:
            current.display_name = req.display_name if req.display_name.strip() else None
        if req.email is not None:
            current.email = req.email if req.email.strip() else None
        if req.phone is not None:
            current.phone = req.phone if req.phone.strip() else None
        if req.department is not None:
            current.department = req.department if req.department.strip() else None
        if hasattr(current, "avatar_url") and req.avatar_url is not None:
            current.avatar_url = req.avatar_url if req.avatar_url.strip() else None

        current.updated_at = datetime.now()
        session.commit()
        return {"success": True, "message": "个人信息更新成功"}
    except IntegrityError as e:
        session.rollback()
        error_msg = str(e.orig) if hasattr(e, 'orig') else str(e)
        if "UNIQUE constraint failed" in error_msg:
            if "email" in error_msg:
                raise HTTPException(status_code=400, detail="该邮箱已被其他用户使用")
            raise HTTPException(status_code=400, detail="数据冲突，请检查输入信息是否重复")
        raise HTTPException(status_code=400, detail=f"数据库约束错误: {error_msg}")
    except SQLAlchemyError as e:
        session.rollback()
        raise HTTPException(status_code=500, detail=f"数据库操作失败: {str(e)}")
    except HTTPException:
        raise
    except Exception as e:
        session.rollback()
        raise HTTPException(status_code=500, detail=f"服务器内部错误: {str(e)}")
    finally:
        session.close()


class ChangePasswordRequest(BaseModel):
    current_password: str = Field(..., min_length=1)
    new_password: str = Field(..., min_length=6)


@app.post("/api/profile/password")
def change_profile_password(token: str = Query(...), req: ChangePasswordRequest = None):
    """修改当前用户密码"""
    if req is None:
        raise HTTPException(status_code=400, detail="请求体不能为空")
    session = get_session()
    try:
        current = get_current_user(session, token)
        if not current:
            raise HTTPException(status_code=401, detail="未登录")

        # 验证当前密码
        if not verify_password(req.current_password, current.password_hash):
            raise HTTPException(status_code=400, detail="当前密码错误")

        # 密码强度校验
        pwd = req.new_password
        if len(pwd) < 6:
            raise HTTPException(status_code=400, detail="密码长度至少6位")
        if not any(c.isalpha() for c in pwd):
            raise HTTPException(status_code=400, detail="密码需包含字母")
        if not any(c.isdigit() for c in pwd):
            raise HTTPException(status_code=400, detail="密码需包含数字")

        # 更新密码
        current.password_hash = hash_password(pwd)
        current.updated_at = datetime.now()
        session.commit()
        return {"success": True, "message": "密码修改成功"}
    finally:
        session.close()


@app.post("/api/profile/avatar")
async def upload_profile_avatar(token: str = Query(...), avatar: UploadFile = File(...)):
    """上传用户头像"""
    session = get_session()
    try:
        current = get_current_user(session, token)
        if not current:
            raise HTTPException(status_code=401, detail="未登录")

        # 验证文件类型
        if not avatar.content_type or not avatar.content_type.startswith("image/"):
            raise HTTPException(status_code=400, detail="请上传图片文件")

        # 读取文件内容并验证大小
        content = await avatar.read()
        if len(content) > 5 * 1024 * 1024:  # 5MB
            raise HTTPException(status_code=400, detail="图片大小不能超过5MB")

        # 生成唯一文件名
        file_extension = os.path.splitext(avatar.filename or "avatar.jpg")[1] or ".jpg"
        unique_filename = f"{uuid.uuid4().hex}{file_extension}"
        file_path = AVATAR_DIR / unique_filename

        # 保存文件
        with open(file_path, "wb") as f:
            f.write(content)

        # 删除旧头像（如果存在）
        if current.avatar_url:
            try:
                old_filename = os.path.basename(current.avatar_url)
                old_file_path = AVATAR_DIR / old_filename
                if old_file_path.exists():
                    old_file_path.unlink()
            except:
                pass

        # 生成头像URL
        avatar_url = f"http://localhost:8000/uploads/avatars/{unique_filename}"
        
        # 更新用户头像
        current.avatar_url = avatar_url
        current.updated_at = datetime.now()
        session.commit()

        return {"success": True, "avatar_url": avatar_url}
    finally:
        session.close()


@app.post("/api/products/search")
def products_search(req: ProductSearchRequest, token: str = Query(...)):
    """产品搜索（需要登录），返回包含双版本单位信息"""
    session = get_session()
    try:
        current = get_current_user(session, token)
        if not current or not has_permission(current, Permission.PRODUCT_VIEW):
            raise HTTPException(status_code=403, detail="您的权限不足，无法执行此操作")
        if not req.keyword:
            from app.models.models import Product
            products = session.query(Product).limit(50).all()
        else:
            products = search_products(session, req.keyword)
        return {
            "products": [
                {
                    "id": p.id,
                    "sku": p.sku,
                    "name": p.name,
                    "model": p.model,
                    "specification": p.specification,
                    "price": p.price,
                    "stock_quantity": p.stock_quantity,
                    "length_cm": p.length_cm,
                    "width_cm": p.width_cm,
                    "height_cm": p.height_cm,
                    "gross_weight_kg": p.gross_weight_kg,
                    "length_inch": p.length_inch,
                    "width_inch": p.width_inch,
                    "height_inch": p.height_inch,
                    "gross_weight_lb": p.gross_weight_lb,
                    "unit_converted": p.unit_converted,
                    "converted_at": p.converted_at.isoformat() if p.converted_at else None,
                    "category": p.category,
                    "brand": p.brand,
                    "supplier": p.supplier,
                    "description": p.description,
                    "status": p.status,
                    "created_at": p.created_at.isoformat() if p.created_at else None,
                    "updated_at": p.updated_at.isoformat() if p.updated_at else None,
                }
                for p in products
            ]
        }
    finally:
        session.close()


# ========== 产品添加 API ==========

@app.post("/api/products")
def products_create(req: ProductCreateRequest, token: str = Query(...)):
    """手动添加产品（需要 product.create 权限），自动进行单位转换"""
    session = get_session()
    try:
        current = get_current_user(session, token)
        if not current or not has_permission(current, Permission.PRODUCT_CREATE):
            raise HTTPException(status_code=403, detail="您的权限不足，无法添加产品")

        data = req.model_dump()
        product, errors = create_product(session, data, auto_convert=True)
        if not product:
            raise HTTPException(status_code=400, detail={"message": "产品添加失败", "errors": errors})

        return {
            "success": True,
            "message": "产品添加成功",
            "product": {
                "id": product.id,
                "sku": product.sku,
                "name": product.name,
                "model": product.model,
                "specification": product.specification,
                "price": product.price,
                "stock_quantity": product.stock_quantity,
                "length_cm": product.length_cm,
                "width_cm": product.width_cm,
                "height_cm": product.height_cm,
                "gross_weight_kg": product.gross_weight_kg,
                "length_inch": product.length_inch,
                "width_inch": product.width_inch,
                "height_inch": product.height_inch,
                "gross_weight_lb": product.gross_weight_lb,
                "unit_converted": product.unit_converted,
                "category": product.category,
                "brand": product.brand,
                "supplier": product.supplier,
                "description": product.description,
                "status": product.status,
                "created_at": product.created_at.isoformat() if product.created_at else None,
            }
        }
    finally:
        session.close()


@app.post("/api/products/check-sku")
def products_check_sku(sku: str = Query(...), token: str = Query(...)):
    """检查 SKU 是否已存在"""
    session = get_session()
    try:
        current = get_current_user(session, token)
        if not current:
            raise HTTPException(status_code=401, detail="未登录")

        exists = check_duplicate_sku(session, sku)
        return {"exists": exists, "sku": sku}
    finally:
        session.close()


@app.post("/api/products/import/excel")
async def products_import_excel(file: UploadFile = File(...), token: str = Query(...)):
    """Excel 批量导入产品（需要 product.create 权限）"""
    session = get_session()
    try:
        current = get_current_user(session, token)
        if not current or not has_permission(current, Permission.PRODUCT_CREATE):
            raise HTTPException(status_code=403, detail="您的权限不足，无法导入产品")

        content = await file.read()

        # 解析 Excel
        valid_records, parse_errors = parse_excel_to_products(content, file.filename)

        if not valid_records and parse_errors:
            # 只有解析错误，没有有效数据
            return {
                "success": False,
                "message": "Excel 解析失败",
                "total": 0,
                "success_count": 0,
                "failed_count": len(parse_errors),
                "parse_errors": parse_errors,
                "import_errors": [],
            }

        # 批量创建
        success_count, failed_records = bulk_create_products(session, valid_records)

        return {
            "success": success_count > 0,
            "message": f"导入完成：成功 {success_count} 条，失败 {len(failed_records) + len(parse_errors)} 条",
            "total": len(valid_records) + len(parse_errors),
            "success_count": success_count,
            "failed_count": len(failed_records) + len(parse_errors),
            "parse_errors": parse_errors,
            "import_errors": failed_records,
        }
    finally:
        session.close()


@app.get("/api/products/import/template")
def products_import_template():
    """下载 Excel 导入模板（含填写说明）"""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from fastapi.responses import StreamingResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "产品导入模板"

    # 表头
    headers = [
        "sku", "name", "model", "specification", "price",
        "stock_quantity", "length_cm", "width_cm", "height_cm",
        "gross_weight_kg", "category", "brand", "supplier", "description"
    ]
    ws.append(headers)

    # 示例数据
    example = [
        "SKU001", "示例产品", "Model-A", "10x20x30cm", 99.99,
        100, 10.0, 20.0, 30.0, 1.5, "电子产品", "示例品牌", "示例供应商", "产品描述"
    ]
    ws.append(example)

    # 样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 调整列宽
    column_widths = [15, 20, 15, 20, 10, 15, 12, 12, 12, 15, 12, 12, 18, 25]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    # 添加填写说明工作表
    ws_help = wb.create_sheet("填写说明")
    ws_help.append(["字段名", "说明", "是否必填", "示例"])
    help_data = [
        ["sku", "产品唯一编码", "必填", "SKU001"],
        ["name", "产品名称", "必填", "示例产品"],
        ["model", "产品型号", "选填", "Model-A"],
        ["specification", "规格描述", "选填", "10x20x30cm"],
        ["price", "价格", "选填", "99.99"],
        ["stock_quantity", "库存数量", "选填", "100"],
        ["length_cm", "长度（厘米）", "必填", "10.0"],
        ["width_cm", "宽度（厘米）", "必填", "20.0"],
        ["height_cm", "高度（厘米）", "必填", "30.0"],
        ["gross_weight_kg", "毛重（千克）", "必填", "1.5"],
        ["category", "产品分类", "选填", "电子产品"],
        ["brand", "品牌", "选填", "示例品牌"],
        ["supplier", "供应商", "选填", "示例供应商"],
        ["description", "产品描述", "选填", "产品描述"],
    ]
    for row in help_data:
        ws_help.append(row)

    # 说明表样式
    for cell in ws_help[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for col in ws_help.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws_help.column_dimensions[column].width = min(max_length + 4, 40)

    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=product_import_template.xlsx"}
    )


@app.post("/api/products/convert-all")
def products_convert_all(token: str = Query(...)):
    """批量转换所有未转换产品的单位（需要 product.create 权限）"""
    session = get_session()
    try:
        current = get_current_user(session, token)
        if not current or not has_permission(current, Permission.PRODUCT_CREATE):
            raise HTTPException(status_code=403, detail="您的权限不足，无法执行此操作")

        success_count, errors = convert_all_products(session)
        return {
            "success": True,
            "message": f"单位转换完成：成功 {success_count} 条",
            "converted_count": success_count,
            "errors": errors,
        }
    finally:
        session.close()


@app.get("/api/products/{sku}/detail")
def products_detail(sku: str, token: str = Query(...)):
    """获取产品详细信息（含双版本单位）"""
    session = get_session()
    try:
        current = get_current_user(session, token)
        if not current or not has_permission(current, Permission.PRODUCT_VIEW):
            raise HTTPException(status_code=403, detail="您的权限不足，无法执行此操作")

        product = get_product_with_converted(session, sku)
        if not product:
            raise HTTPException(status_code=404, detail="产品不存在")

        return product
    finally:
        session.close()


# ========== RBAC 权限系统 API ==========

@app.get("/api/roles")
def list_roles(token: str = Query(...)):
    """获取角色列表与权限矩阵（管理员）"""
    session = get_session()
    try:
        current = get_current_user(session, token)
        if not current or not has_permission(current, Permission.ROLE_MANAGE):
            raise HTTPException(status_code=403, detail="您的权限不足，无法执行此操作")

        roles = []
        for role_key in [UserRole.OPERATOR, UserRole.FINANCE, UserRole.ADMIN]:
            roles.append({
                "key": role_key,
                "name": ROLE_DISPLAY_NAMES.get(role_key, role_key),
                "permissions": ROLE_PERMISSIONS.get(role_key, []),
                "permission_names": [
                    PERMISSION_DISPLAY_NAMES.get(p, p)
                    for p in ROLE_PERMISSIONS.get(role_key, [])
                ],
            })

        return {
            "roles": roles,
            "matrix": ROLE_PERMISSION_MATRIX,
        }
    finally:
        session.close()


@app.get("/api/roles/permissions")
def list_all_permissions(token: str = Query(...)):
    """获取所有可用权限列表（管理员）"""
    session = get_session()
    try:
        current = get_current_user(session, token)
        if not current or not has_permission(current, Permission.ROLE_MANAGE):
            raise HTTPException(status_code=403, detail="您的权限不足，无法执行此操作")

        return {
            "permissions": [
                {
                    "key": p,
                    "name": PERMISSION_DISPLAY_NAMES.get(p, p),
                }
                for p in Permission.ALL_PERMISSIONS
            ]
        }
    finally:
        session.close()


@app.post("/api/roles/check")
def check_permission_api(req: PermissionCheckRequest, token: str = Query(...)):
    """检查当前用户是否拥有指定权限"""
    session = get_session()
    try:
        current = get_current_user(session, token)
        if not current:
            raise HTTPException(status_code=401, detail="未登录或登录已过期")

        allowed = has_permission(current, req.permission)
        return {
            "permission": req.permission,
            "granted": allowed,
            "role": current.role,
        }
    finally:
        session.close()


@app.post("/api/users/{user_id}/role")
def assign_user_role(user_id: int, req: RoleAssignRequest, token: str = Query(...)):
    """分配用户角色（仅管理员）"""
    session = get_session()
    try:
        current = get_current_user(session, token)
        if not current or not has_permission(current, Permission.ROLE_ASSIGN):
            raise HTTPException(status_code=403, detail="您的权限不足，无法执行此操作")

        # 验证目标角色是否有效
        if req.role not in [UserRole.OPERATOR, UserRole.FINANCE, UserRole.ADMIN]:
            raise HTTPException(status_code=400, detail="无效的角色")

        # 不能修改自己的角色
        if current.id == user_id:
            raise HTTPException(status_code=400, detail="不能修改自己的角色")

        user = session.query(User).filter(User.id == user_id).first()
        if not user:
            raise HTTPException(status_code=404, detail="用户不存在")

        # 记录旧角色
        old_role = user.role
        user.role = req.role
        user.updated_at = datetime.now()
        session.commit()

        return {
            "success": True,
            "message": f"角色已更新: {ROLE_DISPLAY_NAMES.get(old_role, old_role)} -> {ROLE_DISPLAY_NAMES.get(req.role, req.role)}",
            "user_id": user.id,
            "old_role": old_role,
            "new_role": req.role,
        }
    finally:
        session.close()


@app.get("/api/users/{user_id}/permissions")
def get_user_permissions(user_id: int, token: str = Query(...)):
    """获取指定用户的权限列表（管理员）"""
    session = get_session()
    try:
        current = get_current_user(session, token)
        if not current or not has_permission(current, Permission.USER_MANAGE):
            raise HTTPException(status_code=403, detail="您的权限不足，无法执行此操作")

        user = session.query(User).filter(User.id == user_id).first()
        if not user:
            raise HTTPException(status_code=404, detail="用户不存在")

        perms = ROLE_PERMISSIONS.get(user.role, [])
        return {
            "user_id": user.id,
            "username": user.username,
            "role": user.role,
            "role_name": ROLE_DISPLAY_NAMES.get(user.role, user.role),
            "permissions": perms,
            "permission_names": [PERMISSION_DISPLAY_NAMES.get(p, p) for p in perms],
        }
    finally:
        session.close()
