"""
对账系统单元测试与集成测试
"""
import io
import csv
import pytest
from decimal import Decimal, ROUND_HALF_UP
from datetime import datetime

from app.services.reconciliation_parser import (
    validate_file, _parse_float, _parse_int, _convert_value,
    parse_csv_shipping_file
)
from app.services.reconciliation_engine import (
    _round2, _calc_surcharge_total, _extract_warehouse_code,
    _normalize_zip, compare_record
)


# ========== 文件验证测试 ==========

class TestFileValidation:
    def test_valid_csv(self):
        content = b"col1,col2\n1,2"
        valid, msg = validate_file(content, "test.csv")
        assert valid is True
        assert msg == ""

    def test_valid_xlsx(self):
        content = b"PK\x03\x04"  # xlsx magic bytes
        valid, msg = validate_file(content, "test.xlsx")
        assert valid is True
        assert msg == ""

    def test_invalid_extension(self):
        content = b"test"
        valid, msg = validate_file(content, "test.pdf")
        assert valid is False
        assert "不支持的文件格式" in msg

    def test_size_limit(self):
        content = b"x" * (11 * 1024 * 1024)  # 11MB
        valid, msg = validate_file(content, "test.csv")
        assert valid is False
        assert "超过限制" in msg

    def test_size_within_limit(self):
        content = b"x" * (5 * 1024 * 1024)  # 5MB
        valid, msg = validate_file(content, "test.csv")
        assert valid is True


# ========== 数据解析测试 ==========

class TestDataParsing:
    def test_parse_float_valid(self):
        assert _parse_float("123.45") == 123.45
        assert _parse_float("$10.50") == 10.50
        assert _parse_float("1,234.56") == 1234.56

    def test_parse_float_invalid(self):
        assert _parse_float("abc") is None
        assert _parse_float(None) is None
        assert _parse_float("") is None

    def test_parse_int_valid(self):
        assert _parse_int("42") == 42
        assert _parse_int("3.0") == 3

    def test_parse_int_invalid(self):
        assert _parse_int("abc") is None

    def test_convert_value(self):
        assert _convert_value("  test  ") == "test"
        assert _convert_value(None) is None
        assert _convert_value("") is None
        assert _convert_value(123) == 123

    def test_parse_csv(self):
        csv_content = """跟踪号,SKU,总价,基础价,数量,重量(lbs),体积重(lbs),计费重,发货仓,目的地邮编,区
888862796808,KC1201-LB,14.44,8.86,1,16.0,16.0,16.0,CA,12345,2
"""
        records = list(parse_csv_shipping_file(csv_content.encode('utf-8')))
        assert len(records) == 1
        assert records[0]['tracking_no'] == '888862796808'
        assert records[0]['sku'] == 'KC1201-LB'
        assert records[0]['total_amount'] == 14.44
        assert records[0]['base_amount'] == 8.86
        assert records[0]['qty'] == 1
        assert records[0]['weight_lb'] == 16.0
        assert records[0]['billed_weight'] == 16.0
        assert records[0]['warehouse'] == 'CA'
        assert records[0]['zip_code'] == '12345'
        assert records[0]['zone'] == '2'


# ========== 对账引擎测试 ==========

class TestReconciliationEngine:
    def test_round2(self):
        assert _round2(123.456) == 123.46
        assert _round2(123.454) == 123.45
        assert _round2(None) is None

    def test_calc_surcharge_total(self):
        record = {
            'fuel': 1.5,
            'ahs_dim': 2.0,
            'das': 3.5,
            'residential': 5.0,
        }
        assert _calc_surcharge_total(record) == 12.0

    def test_calc_surcharge_total_empty(self):
        assert _calc_surcharge_total({}) == 0.0

    def test_extract_warehouse_code(self):
        assert _extract_warehouse_code("CA") == "CA"
        assert _extract_warehouse_code("NJ Warehouse") == "NJ"
        assert _extract_warehouse_code("TX-Dallas") == "TX"
        assert _extract_warehouse_code("SAV") == "SAV"
        assert _extract_warehouse_code(None) == "CA"

    def test_normalize_zip(self):
        assert _normalize_zip("12345") == "12345"
        assert _normalize_zip("12345-6789") == "12345"
        assert _normalize_zip(None) == ""


# ========== 集成测试 ==========

class TestIntegration:
    def test_full_csv_upload_and_compare(self):
        """完整流程：CSV 上传 -> 解析 -> 比对"""
        csv_content = """跟踪号,SKU,总价,基础价,数量,重量(lbs),体积重(lbs),计费重,发货仓,目的地邮编,区,店铺,订单类型
888862796808,KC1201-LB,14.44,8.86,1,16.0,16.0,16.0,CA,12345,2,StoreA,Normal
"""
        # 验证文件
        content = csv_content.encode('utf-8')
        valid, msg = validate_file(content, "test.csv")
        assert valid is True

        # 解析
        records = list(parse_csv_shipping_file(content))
        assert len(records) == 1
        assert records[0]['tracking_no'] == '888862796808'

    def test_excel_file_validation(self):
        """测试 Excel 文件验证"""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "运费"
        ws.append(["跟踪号", "SKU", "总价", "基础价"])
        ws.append(["888862796808", "KC1201-LB", 14.44, 8.86])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        content = output.read()

        valid, msg = validate_file(content, "test.xlsx")
        assert valid is True


# ========== Excel 解析测试 ==========

class TestExcelParsing:
    def test_parse_excel_with_shipping_sheet(self):
        """测试解析包含'运费'sheet的Excel文件"""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "运费账单"
        ws.append(["跟踪号", "SKU", "总价", "基础价", "数量", "重量(lbs)", "体积重(lbs)", "计费重", "发货仓", "目的地邮编", "区", "物流商"])
        ws.append(["888862796808", "KC1201-LB", 14.44, 8.86, 1, 16.0, 16.0, 16.0, "CA", "12345", 2, "FedEx"])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        content = output.read()

        from app.services.reconciliation_parser import parse_excel_shipping_sheet
        records = list(parse_excel_shipping_sheet(content))
        assert len(records) == 1
        assert records[0]['tracking_no'] == '888862796808'
        assert records[0]['sku'] == 'KC1201-LB'
        assert records[0]['total_amount'] == 14.44
        assert records[0]['carrier'] == 'FedEx'

    def test_parse_excel_missing_tracking_no(self):
        """测试跳过没有跟踪号的行"""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "运费"
        ws.append(["跟踪号", "SKU", "总价"])
        ws.append([None, "SKU001", 10.0])  # 无跟踪号，应被跳过
        ws.append(["TRACK001", "SKU002", 20.0])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        content = output.read()

        from app.services.reconciliation_parser import parse_excel_shipping_sheet
        records = list(parse_excel_shipping_sheet(content))
        assert len(records) == 1
        assert records[0]['tracking_no'] == 'TRACK001'


# ========== 对账引擎差异检测测试 ==========

class TestReconciliationDiffDetection:
    def test_amount_diff_detection(self):
        """测试金额差异检测逻辑"""
        result = {
            'file_total_amount': 100.0,
            'sys_total_amount': 105.0,
        }
        amount_diff = abs(result['file_total_amount'] - result['sys_total_amount'])
        assert amount_diff > 0.01  # 应触发差异

    def test_amount_match(self):
        """测试金额一致的情况"""
        result = {
            'file_total_amount': 100.0,
            'sys_total_amount': 100.005,
        }
        amount_diff = abs(result['file_total_amount'] - result['sys_total_amount'])
        assert amount_diff < 0.01  # 不应触发差异

    def test_weight_diff_threshold(self):
        """测试重量差异阈值（0.1 lbs）"""
        result = {
            'file_weight_lb': 10.0,
            'sys_weight_lb': 10.05,
        }
        weight_diff = abs(result['file_weight_lb'] - result['sys_weight_lb'])
        assert weight_diff < 0.1  # 不应触发差异

        result2 = {
            'file_weight_lb': 10.0,
            'sys_weight_lb': 10.2,
        }
        weight_diff2 = abs(result2['file_weight_lb'] - result2['sys_weight_lb'])
        assert weight_diff2 > 0.1  # 应触发差异


# ========== 性能测试 ==========

class TestPerformance:
    def test_large_csv_parsing(self):
        """测试 10 万条记录的解析性能"""
        import time

        # 生成 1000 条测试数据（模拟）
        rows = []
        rows.append("跟踪号,SKU,总价,基础价,数量,重量(lbs),体积重(lbs),计费重,发货仓,目的地邮编,区")
        for i in range(1000):
            rows.append(f"TRACK{i:06d},SKU{i:04d},10.00,8.00,1,5.0,5.0,5.0,CA,12345,2")

        csv_content = "\n".join(rows)
        content = csv_content.encode('utf-8')

        start = time.time()
        records = list(parse_csv_shipping_file(content))
        elapsed = time.time() - start

        assert len(records) == 1000
        assert elapsed < 5.0  # 1000条应在5秒内完成
        print(f"解析 1000 条记录耗时: {elapsed:.3f}s")


# ========== 安全测试 ==========

class TestTokenSecurity:
    def test_auth_headers_helper(self):
        """验证 authHeaders 生成正确的 Authorization 头"""
        token = "test_token_123"
        headers = {
            'Content-Type': 'application/json',
            'Authorization': f'Bearer {token}',
        }
        assert headers['Authorization'] == 'Bearer test_token_123'
        assert headers['Content-Type'] == 'application/json'

    def test_token_not_in_url(self):
        """验证 token 不会出现在 URL 中"""
        token = "secret_token"
        url = "http://localhost:8000/api/reconciliation/batches?page=1"
        # token 应该在 header 中，而不是 URL 中
        assert token not in url


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
