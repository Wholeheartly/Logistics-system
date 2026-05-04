"""
数据导入脚本 — 从 Excel 文件中抽取数据并写入数据库。
运行: cd backend && python scripts/seed_database.py
  本地: python scripts/seed_database.py
  生产: DATABASE_URL=postgresql://... python scripts/seed_database.py
"""
import os, sys
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import openpyxl
from sqlalchemy import create_engine
from sqlalchemy.orm import Session

from app.models.models import Base, Carrier, Product, BaseRate, SurchargeConfig, ZoneMapping, UnserviceableZip

DB_URL = os.environ.get("DATABASE_URL", "")
if not DB_URL:
    PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
    DB_PATH = os.path.join(PROJECT_ROOT, "backend", "logistics.db")
    DB_URL = f"sqlite:///{DB_PATH}"
    DATA_DIR = PROJECT_ROOT
else:
    DATA_DIR = os.environ.get("DATA_DIR", os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")))

engine = create_engine(DB_URL, echo=False)


def create_db():
    Base.metadata.create_all(engine)
    print(f"数据库已创建: {DB_URL.split('@')[-1] if '@' in DB_URL else DB_URL}")


def import_products(session: Session):
    """导入产品数据库"""
    path = os.path.join(DATA_DIR, "产品数据库.xlsx")
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        sku, length, width, height, weight = row[0], row[1], row[2], row[3], row[4]
        if sku is None:
            continue
        session.add(Product(
            sku=str(sku).strip(),
            length_cm=float(length) if length else 0,
            width_cm=float(width) if width else 0,
            height_cm=float(height) if height else 0,
            gross_weight_kg=float(weight) if weight else 0,
        ))
        count += 1
    session.commit()
    print(f"产品导入完成: {count} 条")


def import_carriers(session: Session):
    """导入五大物流渠道配置（数据来源：系统提示词）"""
    carriers = [
        Carrier(name="商仓 FedEx Ground&HD", dim_factor=270, residential_fee=3.00, fuel_rate=0.20,
                max_weight_lb=150, has_residential_fee=True, transit_time="1-5个工作日"),
        Carrier(name="商仓 FDX Ground Economy", dim_factor=139, residential_fee=0.00, fuel_rate=0.20,
                max_weight_lb=35, has_residential_fee=False, transit_time="1-5个工作日"),
        Carrier(name="商仓 Amazon Shipping", dim_factor=300, residential_fee=0.00, fuel_rate=0.20,
                max_weight_lb=150, has_residential_fee=False, transit_time="1-5个工作日"),
        Carrier(name="Fimile Ground", dim_factor=250, residential_fee=0.00, fuel_rate=0.20,
                max_weight_lb=150, has_residential_fee=False, transit_time="1-5个工作日"),
        Carrier(name="GOFO Ground", dim_factor=250, residential_fee=3.00, fuel_rate=0.20,
                max_weight_lb=150, has_residential_fee=True, transit_time="1-5个工作日"),
    ]
    for c in carriers:
        session.add(c)
    session.commit()
    print(f"物流渠道导入完成: {len(carriers)} 条")


def import_base_rates_from_excel(session: Session, carrier_name: str, file_name: str, sheet_name: str):
    """从 Excel 的费率 sheet 导入基础运费"""
    carrier = session.query(Carrier).filter(Carrier.name == carrier_name).first()
    if not carrier:
        print(f"  找不到渠道: {carrier_name}")
        return

    path = os.path.join(DATA_DIR, file_name)
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name]

    count = 0
    for row in ws.iter_rows(min_row=3, values_only=True):
        # 列结构: [None, weight_lb, weight_kg, zone2, zone3, zone4, zone5, zone6, zone7, zone8, ...]
        if row[1] is None or not isinstance(row[1], (int, float)):
            # 可能到了附加费配置行，跳过
            if row[9] is not None and isinstance(row[9], str):
                continue
            continue
        weight_lb = int(row[1])
        for zone_idx, zone_val in enumerate([2, 3, 4, 5, 6, 7, 8], start=3):
            if zone_idx < len(row) and row[zone_idx] is not None and isinstance(row[zone_idx], (int, float)):
                session.add(BaseRate(
                    carrier_id=carrier.id,
                    weight_lb=weight_lb,
                    zone=zone_val,
                    rate=float(row[zone_idx]),
                ))
                count += 1
    session.commit()
    print(f"  基础运费: {carrier_name} — {count} 条")


def import_surcharge_configs(session: Session):
    """导入附加费配置（数据来源：系统提示词）"""
    carriers = {c.name: c.id for c in session.query(Carrier).all()}

    configs = [
        # === 商仓 FedEx Ground&HD ===
        ("商仓 FedEx Ground&HD", "AHS-Dim", "2", 6.195),
        ("商仓 FedEx Ground&HD", "AHS-Dim", "3-4", 6.8775),
        ("商仓 FedEx Ground&HD", "AHS-Dim", "5-6", 8.148),
        ("商仓 FedEx Ground&HD", "AHS-Dim", "7-8", 8.5575),
        ("商仓 FedEx Ground&HD", "AHS-Weight", "2", 7.82),
        ("商仓 FedEx Ground&HD", "AHS-Weight", "3-4", 8.5425),
        ("商仓 FedEx Ground&HD", "AHS-Weight", "5-6", 9.5625),
        ("商仓 FedEx Ground&HD", "AHS-Weight", "7-8", 9.9875),
        ("商仓 FedEx Ground&HD", "AHS-Packaging", "2", 3.50),
        ("商仓 FedEx Ground&HD", "AHS-Packaging", "3-4", 3.50),
        ("商仓 FedEx Ground&HD", "AHS-Packaging", "5-6", 3.50),
        ("商仓 FedEx Ground&HD", "AHS-Packaging", "7-8", 3.50),
        ("商仓 FedEx Ground&HD", "Oversize", "2", 40.80),
        ("商仓 FedEx Ground&HD", "Oversize", "3-4", 44.00),
        ("商仓 FedEx Ground&HD", "Oversize", "5-6", 51.20),
        ("商仓 FedEx Ground&HD", "Oversize", "7-8", 52.80),
        ("商仓 FedEx Ground&HD", "DAS", "all", 3.96),
        ("商仓 FedEx Ground&HD", "Extended_DAS", "all", 5.28),
        ("商仓 FedEx Ground&HD", "Remote_DAS", "all", 11.70),

        # === 商仓 Amazon Shipping ===
        ("商仓 Amazon Shipping", "AHS-Dim", "2", 29.26),
        ("商仓 Amazon Shipping", "AHS-Dim", "3-4", 32.59),
        ("商仓 Amazon Shipping", "AHS-Dim", "5+", 37.57),
        ("商仓 Amazon Shipping", "AHS-Weight", "2", 45.89),
        ("商仓 Amazon Shipping", "AHS-Weight", "3-4", 49.88),
        ("商仓 Amazon Shipping", "AHS-Weight", "5+", 55.20),
        ("商仓 Amazon Shipping", "AHS-NonStandard", "2", 25.94),
        ("商仓 Amazon Shipping", "AHS-NonStandard", "3-4", 30.59),
        ("商仓 Amazon Shipping", "AHS-NonStandard", "5+", 32.59),
        ("商仓 Amazon Shipping", "Oversize", "2", 240.00),
        ("商仓 Amazon Shipping", "Oversize", "3-4", 260.00),
        ("商仓 Amazon Shipping", "Oversize", "5+", 290.00),
        ("商仓 Amazon Shipping", "NonStandard_Fee", "2", 11.00),
        ("商仓 Amazon Shipping", "NonStandard_Fee", "3-4", 12.25),
        ("商仓 Amazon Shipping", "NonStandard_Fee", "5+", 14.15),
        ("商仓 Amazon Shipping", "DAS", "all", 4.45),
        ("商仓 Amazon Shipping", "Extended_DAS", "all", 5.55),
        ("商仓 Amazon Shipping", "Remote_DAS", "all", 16.75),

        # === 商仓 FDX Ground Economy ===
        ("商仓 FDX Ground Economy", "DAR", "all", 0.70),
        ("商仓 FDX Ground Economy", "USPS_NonMach", "all", 6.50),
        ("商仓 FDX Ground Economy", "DAS", "all", 4.50),
        ("商仓 FDX Ground Economy", "Extended_DAS", "all", 5.50),
        ("商仓 FDX Ground Economy", "Remote_DAS", "all", 15.50),

        # === Fimile Ground ===
        ("Fimile Ground", "AHS-Dim", "2", 4.95),
        ("Fimile Ground", "AHS-Dim", "3-4", 5.28),
        ("Fimile Ground", "AHS-Dim", "5-6", 5.85),
        ("Fimile Ground", "AHS-Dim", "7-8", 6.08),
        ("Fimile Ground", "AHS-Weight", "2", 5.50),
        ("Fimile Ground", "AHS-Weight", "3-4", 5.50),
        ("Fimile Ground", "AHS-Weight", "5-6", 6.15),
        ("Fimile Ground", "AHS-Weight", "7-8", 6.55),
        ("Fimile Ground", "AHS-Packaging", "2", 4.23),
        ("Fimile Ground", "AHS-Packaging", "3-4", 5.26),
        ("Fimile Ground", "AHS-Packaging", "5-6", 5.64),
        ("Fimile Ground", "AHS-Packaging", "7-8", 5.90),
        ("Fimile Ground", "Oversize", "2", 23.00),
        ("Fimile Ground", "Oversize", "3-4", 23.00),
        ("Fimile Ground", "Oversize", "5-6", 25.00),
        ("Fimile Ground", "Oversize", "7-8", 28.00),
        ("Fimile Ground", "DAS", "all", 3.96),
        ("Fimile Ground", "Extended_DAS", "all", 5.28),
        ("Fimile Ground", "Remote_DAS", "all", 11.70),

        # === GOFO Ground ===
        ("GOFO Ground", "AHS-Dim", "2", 5.20),
        ("GOFO Ground", "AHS-Dim", "3-4", 5.68),
        ("GOFO Ground", "AHS-Dim", "5-6", 7.05),
        ("GOFO Ground", "AHS-Dim", "7-8", 7.38),
        ("GOFO Ground", "AHS-Weight", "2", 6.78),
        ("GOFO Ground", "AHS-Weight", "3-4", 7.46),
        ("GOFO Ground", "AHS-Weight", "5-6", 8.56),
        ("GOFO Ground", "AHS-Weight", "7-8", 8.89),
        ("GOFO Ground", "AHS-Packaging", "2", 4.23),
        ("GOFO Ground", "AHS-Packaging", "3-4", 5.26),
        ("GOFO Ground", "AHS-Packaging", "5-6", 5.64),
        ("GOFO Ground", "AHS-Packaging", "7-8", 5.90),
        ("GOFO Ground", "Oversize", "2", 26.68),
        ("GOFO Ground", "Oversize", "3-4", 28.60),
        ("GOFO Ground", "Oversize", "5-6", 30.04),
        ("GOFO Ground", "Oversize", "7-8", 32.96),
        ("GOFO Ground", "DAS", "all", 3.96),
        ("GOFO Ground", "Extended_DAS", "all", 5.28),
    ]

    for name, s_type, z_group, amt in configs:
        cid = carriers.get(name)
        if cid:
            session.add(SurchargeConfig(carrier_id=cid, surcharge_type=s_type, zone_group=z_group, amount=amt))
    session.commit()
    print(f"附加费配置导入完成: {len(configs)} 条")


def import_zone_mappings_shangcang(session: Session):
    """导入商仓的邮编分区表"""
    path = os.path.join(DATA_DIR, "商仓物流2026报价.xlsx")
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["邮编分区表"]

    warehouses = ["CA", "NJ", "SC", "TX"]
    col_map = {0: 2, 1: 4, 2: 6, 3: 8}  # warehouse_idx -> column in Excel

    count = 0
    for row in ws.iter_rows(min_row=8, values_only=True):  # 数据从第8行开始
        zip_prefix = row[1]
        if zip_prefix is None:
            continue
        zip_prefix = str(zip_prefix).strip()[:3]
        for wh_idx, warehouse in enumerate(warehouses):
            col = col_map[wh_idx]
            zone_val = row[col] if col < len(row) else None
            if zone_val is not None:
                session.add(ZoneMapping(warehouse=warehouse, zip_prefix=zip_prefix, zone=int(zone_val)))
                count += 1
    session.commit()
    print(f"商仓邮编分区表导入完成: {count} 条")


def import_zone_mappings_gofo(session: Session):
    """导入 GOFO Ground 的邮编分区表"""
    path = os.path.join(DATA_DIR, "GOFO  Ground 20260415.xlsx")
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["可达邮编"]

    # 列结构: 序号, 邮编, 县市, 省州, 城市, CA_zone, NJ_zone, SAV_zone, 是否偏远
    warehouses = ["CA", "NJ", "SAV"]
    col_map = [5, 6, 7]

    count = 0
    for row in ws.iter_rows(min_row=4, values_only=True):
        zip_code = row[1]
        if zip_code is None:
            continue
        zip_prefix = str(int(zip_code)).zfill(5)[:3]
        for wh_idx, warehouse in enumerate(warehouses):
            col = col_map[wh_idx]
            zone_val = row[col] if col < len(row) else None
            if zone_val is not None:
                session.add(ZoneMapping(warehouse=warehouse, zip_prefix=zip_prefix, zone=int(zone_val)))
                count += 1
    session.commit()
    print(f"GOFO 邮编分区表导入完成: {count} 条")


def import_zone_mappings_fimile(session: Session):
    """导入 Fimile 的邮编分区表（12 个列组，每组 4 列: zip, state, city, zone）"""
    path = os.path.join(DATA_DIR, "Fimile-20260424.xlsx")
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.worksheets[1]

    # 12 个列组的起始列（1-indexed）和对应的仓库
    # Row 2 的 group headers 标明了每条线路所属仓库
    group_starts = [
        (2,  "CA"),   # CA仓发-本地方向 (491 zips)
        (7,  "CA"),   # CA-SFO线 (311 zips)
        (12, "CA"),   # CA-TX线 (179 zips)
        (17, "CA"),   # CA-LAS线 (60 zips)
        (22, "CA"),   # CA-PHX线 (141 zips)
        (27, "CA"),   # CA-PA-NJ-NY线 (971 zips)
        (32, "CA"),   # CA-DMV专线 (255 zips)
        (37, "NJ"),   # NJ仓发-本地方向 PA-NJ-NY (971 zips)
        (42, "NJ"),   # NJ-DMV专线 (255 zips)
        (47, "TX"),   # TX仓发-本地方向 (179 zips)
        (53, "SAV"),  # SAV-FL线 (449 zips)
        (58, "SAV"),  # SAV-ATL线 (140 zips)
    ]

    count = 0
    seen = set()
    for row in ws.iter_rows(min_row=4, values_only=True):
        for start_col_1, warehouse in group_starts:
            ci = start_col_1 - 1  # 转为 0-indexed
            if ci + 3 >= len(row):
                continue
            zip_val = row[ci]       # zip (同列在 row 3)
            zone_val = row[ci + 3]  # zone
            if zip_val is None or zone_val is None:
                continue
            try:
                zip_prefix = str(int(zip_val)).zfill(5)[:3]
                zone_int = int(zone_val)
            except (ValueError, TypeError):
                continue
            key = (warehouse, zip_prefix)
            if key not in seen:
                seen.add(key)
                session.add(ZoneMapping(warehouse=warehouse, zip_prefix=zip_prefix, zone=zone_int))
                count += 1
    session.commit()
    print(f"Fimile 邮编分区表导入完成: {count} 条")


def import_unserviceable_zips(session: Session):
    """导入不可达邮编（AK, HI, PR 等）"""
    states = [
        ("AK", "Alaska"),
        ("HI", "Hawaii"),
        ("PR", "Puerto Rico"),
        ("GU", "Guam"),
        ("AA", "APO/FPO - Americas"),
        ("AE", "APO/FPO - Europe"),
        ("AP", "APO/FPO - Pacific"),
    ]
    for code, desc in states:
        session.add(UnserviceableZip(state_code=code, description=desc))
    session.commit()
    print(f"不可达邮编导入: {len(states)} 条")


def main():
    create_db()
    with Session(engine) as session:
        import_products(session)
        import_carriers(session)
        print("\n导入基础运费表...")
        import_base_rates_from_excel(session, "商仓 FedEx Ground&HD", "商仓物流2026报价.xlsx", "Fedex Ground&HD")
        import_base_rates_from_excel(session, "商仓 FDX Ground Economy", "商仓物流2026报价.xlsx", "FDX Ground Economy")
        import_base_rates_from_excel(session, "商仓 Amazon Shipping", "商仓物流2026报价.xlsx", "Amazon Shipping")
        import_base_rates_from_excel(session, "Fimile Ground", "Fimile-20260424.xlsx", "Fimile-Ground")
        import_base_rates_from_excel(session, "GOFO Ground", "GOFO  Ground 20260415.xlsx", "GOFO-Ground")
        print("\n导入附加费配置...")
        import_surcharge_configs(session)
        print("\n导入邮编分区表...")
        import_zone_mappings_shangcang(session)
        import_zone_mappings_gofo(session)
        import_zone_mappings_fimile(session)
        print("\n导入不可达邮编...")
        import_unserviceable_zips(session)
        session.commit()
    print("\n所有数据导入完成!")


if __name__ == "__main__":
    main()
